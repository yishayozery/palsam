"""
Extract equipment data from קובץ קיפולים Excel to JSON for DB loading.
Outputs: 4/equipment-data.json
"""
import openpyxl, sys, io, json, re

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('4/קובץ קיפולים 2026 עדכני 3.2.26.xlsx', read_only=True, data_only=True)

result = {
    "companies": [],  # each: { sheet, companyName, locations: [{ name, items: [{ name, qty, association }] }] }
    "vehicles": [],   # each: { type, serialNumber, notes, companyName }
}

def parse_qty(val):
    """Extract numeric quantity from cell value"""
    if val is None or val == '':
        return 0
    s = str(val).strip()
    m = re.match(r'^(\d+)', s)
    if m:
        return int(m.group(1))
    return 0

def classify(identity_str):
    """Determine MILITARY or DONATION from identity column"""
    s = str(identity_str).strip().lower() if identity_str else ''
    if s in ('צהל', 'צבאי', 'צביא'):
        return 'MILITARY'
    if s in ('תרומה',):
        return 'DONATION_COMPANY'
    return 'MILITARY'  # default

# ============ פלה"ק ============
ws = wb['פלהק']
rows = list(ws.iter_rows(values_only=True))

loc_machsan = {"name": "ציוד מחסן", "items": []}
loc_nagmash = {"name": "ציוד נגמשים", "items": []}
loc_riau = {"name": "ציוד ריאו/משא", "items": []}
loc_trumot = {"name": "תרומות פלה\"ק", "items": []}

for i, row in enumerate(rows):
    if i < 3: continue  # skip headers
    vals = [str(c).strip() if c else '' for c in row[:10]]

    # Column A-B: ציוד מחסן
    if vals[0] and vals[0] != 'שם הפריט' and vals[0] != 'תרומות':
        qty = parse_qty(vals[1])
        if qty > 0:
            # Check if we're in the donations section (row 29+)
            if i >= 28:  # 0-indexed, row 29 = index 28
                loc_trumot["items"].append({"name": vals[0], "qty": qty, "association": "DONATION_COMPANY"})
            else:
                loc_machsan["items"].append({"name": vals[0], "qty": qty, "association": "MILITARY"})

    # Column E-F: ציוד נגמשים
    if vals[4] and vals[4] not in ('שם הפריט', 'שם הפריט ', 'ציוד כשיר להפעלה', '77', 'קיטבג מלא'):
        qty = parse_qty(vals[5])
        if qty > 0:
            loc_nagmash["items"].append({"name": vals[4], "qty": qty, "association": "MILITARY"})

    # Column H-I: ציוד ריאו משא
    if vals[7] and vals[7] not in ('שם הפריט',):
        qty = parse_qty(vals[8])
        if qty > 0:
            loc_riau["items"].append({"name": vals[7], "qty": qty, "association": "MILITARY"})

locs_plhk = [l for l in [loc_machsan, loc_nagmash, loc_riau, loc_trumot] if l["items"]]
result["companies"].append({"sheet": "פלהק", "companyName": "פלה\"ק", "locations": locs_plhk})

# ============ פלס"מ ============
ws = wb['פלסמ']
rows = list(ws.iter_rows(values_only=True))

loc_12m = {"name": "מכולה 12 מטר", "items": []}
loc_6m = {"name": "מכולה 6 מטר", "items": []}
loc_general = {"name": "מחסן כללי", "items": []}

for i, row in enumerate(rows):
    if i < 2: continue
    vals = [str(c).strip() if c else '' for c in row[:10]]

    # Columns A-C: מכולה 12 מטר
    if vals[0] and vals[0] not in ('שם פריט',):
        qty = parse_qty(vals[2])
        assoc = classify(vals[1])
        if qty > 0:
            loc_12m["items"].append({"name": vals[0], "qty": qty, "association": assoc})

    # Columns E-G: מכולה 6 מטר
    if vals[4] and vals[4] not in ('פריט', 'פריט '):
        qty = parse_qty(vals[6])
        assoc = classify(vals[5])
        if qty > 0:
            loc_6m["items"].append({"name": vals[4], "qty": qty, "association": assoc})

    # Columns I-J: מחסן כללי
    if vals[8] and vals[8] not in ('שם פריט',):
        qty = parse_qty(vals[9])
        if qty > 0:
            loc_general["items"].append({"name": vals[8], "qty": qty, "association": "MILITARY"})

locs_plsm = [l for l in [loc_12m, loc_6m, loc_general] if l["items"]]
result["companies"].append({"sheet": "פלסמ", "companyName": "פלגת השהייה", "locations": locs_plsm})

# ============ טנ"א ============
ws = wb['טנא']
rows = list(ws.iter_rows(values_only=True))

loc_tna = {"name": "מחסן טנ\"א", "items": []}

for i, row in enumerate(rows):
    if i < 1: continue
    vals = [str(c).strip() if c else '' for c in row[:4]]
    name = vals[1] if vals[1] else ''
    if name and name not in ('פריט', 'קטגוריה'):
        qty = parse_qty(vals[2])
        if qty > 0:
            loc_tna["items"].append({"name": name, "qty": qty, "association": "MILITARY"})

result["companies"].append({"sheet": "טנא", "companyName": "טנא", "locations": [loc_tna]})

# ============ שינוע ============
ws = wb['שינוע']
rows = list(ws.iter_rows(values_only=True))

loc_shinua_personal = {"name": "ציוד אישי שינוע", "items": []}

for i, row in enumerate(rows):
    if i < 2: continue
    vals = [str(c).strip() if c else '' for c in row[:8]]

    # Columns A-D: Vehicles
    serial = vals[2]  # ז"צ
    vtype = vals[3]    # סוג רכב
    notes = vals[0]    # הערות
    if serial and re.match(r'^\d{5,}$', serial):
        vtype_clean = vtype.strip().lower()
        if 'fmtv' in vtype_clean and 'ישנה' in vtype_clean:
            vtype_name = 'FMTV ישנה'
        elif 'fmtv' in vtype_clean and 'מכולה' in vtype_clean:
            vtype_name = 'FMTV למכולה'
        elif 'fmtv' in vtype_clean:
            vtype_name = 'FMTV'
        elif 'אושקוש' in vtype_clean:
            vtype_name = 'אושקוש'
        elif 'ריו' in vtype_clean or 'רין' in vtype_clean:
            vtype_name = 'ריו'
        else:
            vtype_name = vtype.strip()
        result["vehicles"].append({
            "type": vtype_name,
            "serialNumber": serial,
            "notes": notes if notes else None,
            "companyName": "שינוע",
        })

    # Columns F-H: ציוד אישי
    item_name = vals[7]
    item_qty = parse_qty(vals[6])
    if item_name and item_name not in ('ציוד',) and item_qty > 0:
        loc_shinua_personal["items"].append({"name": item_name, "qty": item_qty, "association": "MILITARY"})

if loc_shinua_personal["items"]:
    result["companies"].append({"sheet": "שינוע", "companyName": "שינוע", "locations": [loc_shinua_personal]})

# ============ פת"ן ============
ws = wb['פתן']
rows = list(ws.iter_rows(values_only=True))

loc_gray = {"name": "מכולה אפורה", "items": []}
loc_blue = {"name": "מכולה כחולה", "items": []}
loc_red = {"name": "מכולה אדומה", "items": []}

for i, row in enumerate(rows):
    if i < 2: continue
    vals = [str(c).strip() if c else '' for c in row[:14]]

    # Columns A-D: מכולה אפורה
    if vals[0] and vals[0] not in ('פריט',):
        qty = parse_qty(vals[1])
        assoc = classify(vals[3])
        if qty > 0:
            loc_gray["items"].append({"name": vals[0], "qty": qty, "association": assoc})

    # Columns F-I: מכולה כחולה (skip דולב headers)
    if vals[5] and vals[5] not in ('פריט',) and not vals[5].startswith('דולב'):
        qty = parse_qty(vals[6])
        assoc = classify(vals[8])
        if qty > 0:
            loc_blue["items"].append({"name": vals[5], "qty": qty, "association": assoc})

    # Columns K-N: מכולה אדומה
    if vals[10] and vals[10] not in ('פריט',):
        qty = parse_qty(vals[11])
        assoc = classify(vals[13])
        if qty > 0:
            loc_red["items"].append({"name": vals[10], "qty": qty, "association": assoc})

locs_patan = [l for l in [loc_gray, loc_blue, loc_red] if l["items"]]
result["companies"].append({"sheet": "פתן", "companyName": "פת\"ן", "locations": locs_patan})

wb.close()

# Stats
total_items = sum(len(item) for c in result["companies"] for loc in c["locations"] for item in [loc["items"]])
total_unique = len(set(item["name"] for c in result["companies"] for loc in c["locations"] for item in loc["items"]))
print(f"Companies: {len(result['companies'])}")
for c in result["companies"]:
    items_count = sum(len(loc["items"]) for loc in c["locations"])
    locs = ', '.join(f'{l["name"]}({len(l["items"])})' for l in c["locations"])
    print(f"  {c['companyName']}: {items_count} items in {len(c['locations'])} locations: {locs}")
print(f"Vehicles: {len(result['vehicles'])}")
for v in result["vehicles"]:
    print(f"  {v['type']} ז\"צ {v['serialNumber']}{' — ' + v['notes'] if v['notes'] else ''}")
print(f"Total item rows: {total_items}")

with open('4/equipment-data.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)
print("\nSaved to 4/equipment-data.json")
